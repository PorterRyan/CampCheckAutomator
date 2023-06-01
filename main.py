#!/usr/bin/python3
#
# Automation for camp check spreadsheets at Portola Redwoods State Park
#
# Created with Python 3.9.6
#
# Version 1.1.1.0
# 
# Created by Ryan Porter (github.com/PorterRyan). 
# Copyright 2021, 2023 Ryan Porter. This software is licensed under the GNU 
# General Public License version 3.0. Please see the COPYING file and copy
# of the GNU GPL v3.0 included with this software.
#
# This program is free software: you can redistribute it and/or modify it
# under the terms of the GNU General Public License as published by the
# Free Software Foundation, either version 3 of the License, or (at your
# option) any later version.
#
# This program is distributed in the hope that it will be useful, but
# WITHOUT ANY WARRANTY; without even the implied warranty of
# MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE. See the GNU General
# Public License for more details.
#
# You should have received a copy of the GNU General Public License along
# with this program. If not, see https://www.gnu.org/license/.

import openpyxl, shutil, os, re, logging, sys
from datetime import date, time, datetime, timedelta
from xls2xlsx import XLS2XLSX
from openpyxl.workbook import Workbook
from openpyxl.reader.excel import load_workbook, InvalidFileException
import PySimpleGUI as sg

# Change logging level to DEBUG for increased logging
logging.basicConfig(filename="debug.log", filemode='w', level=logging.INFO)
logging.info(date.today())

print("Camp Check Automator by Ryan Porter")
print("""\nCopyright (c) 2021, 2023 Ryan Porter""")
print("""
        This program is free software: You can redistribute it and/or modify
        it under the terms of the GNU General Public License as published by
        the Free Software Foundation, either version 3 of the License, or
        (at your option) any later version.

        This program is distributed in the hope that it will be useful, but
        WITHOUT ANY WARRANTY; without even the implied warranty of 
        MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE. See the GNU
        General Public License for more details.

        You should have received a copy of the GNU General Public License
        along with this program. If not, see https://www.gnu.org/licenses.

        Please note that this program uses Python modules that are covered
        under their own licenses. The license for Python itself, as well as
        its modules datetime, os, re, logging, sys, and shutil, can be found
        at https://docs.python.org/3/license.html. openpyxl is licensed 
        under the MIT/Expat license. XLS2XLSX is licensed under the MIT
        license. PySimpleGUI is licensed under the GNU Lesser General Public
        License (LGPL3) +. This software may have been compiled for use on
        Microsoft Windows systems using Nuitka, which is licensed under the
        Apache-2.0 license.

        You can contact the author of this program at github.com/PorterRyan.
        """)
print()

#input("Press Enter when you have downloaded a current Due-In Report.")
#print()

def end_date():
  this_day = date.today()
  min_this_day = datetime.combine(this_day, time.min)
  return min_this_day

def remove_old_reservations():      
  for row in current_sheet.iter_rows(min_row=6, min_col=3, max_col=3, max_row=86): 
    # MAX_ROW is the last row with license information in the camp sheet
    # FUTURE: Group camps need their own functions because they are multiple
    # rows long.
    for cell in row:
      logging.debug("Cell value: " + str(cell.internal_value))
      old_date = cell.internal_value
      logging.debug("old_date type is " + str(type(old_date)))
      
      if str(type(old_date)) == "<class 'datetime.datetime'>" and old_date <= end_date():
        logging.debug(cell.internal_value)
        logging.debug(cell.coordinate)
        logging.debug(cell.row)
        reference = "\"B" + str(cell.row) + ":G" + str(cell.row) + "\""
        cell_row = cell.row
        logging.debug("Reference range: " + reference)
        for col in current_sheet.iter_cols(min_row=cell_row, min_col=2, max_col=7, max_row=cell_row):
          for cell in col:
            debug_print = str(current_sheet[cell.coordinate])
            print("Cell " + debug_print.strip("<Cell 'Camp Check'.>") + " cleared.")
            logging.info("Cell " + debug_print.strip("<Cell 'Camp Check'.>") + " cleared.")
            current_sheet[cell.coordinate] = None    

# Load yesterday's camp sheet

# Create Today's date
today = (date.today().strftime('%Y-%m-%d'))
# Create yesterday's date
yesterday = (date.today() - timedelta(days=1)).strftime('%Y-%m-%d')

src = sg.popup_get_file("Select the camp sheet for " + yesterday)
dst = sg.popup_get_file("Save today's camp sheet", save_as = True)

logging.debug("src = " + str(src))
print("Finding latest camp check\n")
logging.info("Finding latest camp check\n")

if src != dst and src != False: # Check if the file already exists
  logging.debug(str(src))
  print("Copying last camp check\n")
  logging.info("Copying last camp check\n")
  print("Creating camp check for " + today + "\n")
  logging.info("Creating camp check for " + today + "\n")
  shutil.copy(src, dst)
  print("Camp check saved as " + dst + "\n")
  logging.info("Camp check saved as " + dst + "\n")
elif src == dst and src != False: # If the file already exists we don't want to accidentally overwrite it.
  print("\n")
  logging.warning("CURRENT CAMP CHECK FILE ALREADY EXISTS!")
  sys.exit("CURRENT CAMP CHECK FILE ALREADY EXISTS!")
  print()
elif src == False: # If there is no prior camp sheet, copy from the template
  print("No prior camp check found, copying from template")
  logging.info("No prior camp check found, copying from template")
  shutil.copy("Camp Check Template.xlsx", dst)
else: # If there is no template or prior sheet, raise error
  raise ValueError("Camp Check Template missing")
  logging.debug("Camp Check Template Missing")

# Open the Spreadsheet
print("Opening camp check")
logging.info("Opening camp check")
camp_check = openpyxl.load_workbook(dst)
logging.debug(camp_check.sheetnames)

current_sheet = camp_check['Camp Check']

# Update camp check header
print("Updating header")
logging.info("Updating header")
current_sheet['A1'] = date.today().strftime('%A %m/%d/%Y').upper()

# Remove Old Reservations First
logging.debug(end_date())
print("Removing old reservations")
remove_old_reservations()

# Convert Due-in Report from XLS to XLSX

print("Converting Due-in Report")
logging.info("Converting Due-in Report")

# User selects the due-in report
latest_due_in = sg.popup_get_file("Select due-in report")

# Convert the xls file to xlsx format
def open_xls_as_xlsx():
   x2x = XLS2XLSX(latest_due_in)
   #x2x.to_xlsx(str(date.today()) + ".xlsx")

   # Save converted due-in report as temp.xlsx
   x2x.to_xlsx(str("temp.xlsx"))

open_xls_as_xlsx()

if os.path.exists("temp.xlsx"):
  print("Due-in Report Converted")
  logging.info("Due-in Report Converted")
else:
  print("Conversion Error!")
  logging.error("Conversion Error!")

#due_in_report = openpyxl.load_workbook("Due-in Reports\\" + str(date.today()) + ".xlsx")
due_in_report = openpyxl.load_workbook("temp.xlsx")
due_in_sheet = due_in_report['DueInReport']

# START SITE NUMBER ORDERING
print("Start Site Number Ordering")
logging.info("Start Site Number Ordering")
unsorted_names = {}
names = {}

for row in due_in_sheet.iter_rows(min_row=3, min_col=1, max_col=1, max_row=63):
  for cell in row:

    if str(cell.data_type) == "s":
      site_row = cell.row
      site_col = cell.column + 3
      site_coord = due_in_sheet.cell(site_row, site_col).value
      full_name = str(cell.value).split(', ')
      name = full_name[0]

      print("Site Number " + str(site_coord) + ": " + str(name))
      logging.info("Site Number " + str(site_coord) + ": " + str(name))

      unsorted_names[str(site_coord)] = str(name)

print()

names = dict(sorted(unsorted_names.items(), key=lambda item: item[0]))

logging.debug(str(names))
logging.debug(str(list(names)))


# END SITE NUMBER ORDERING
print("Site Number Ordering Finished")
logging.info("Site Number Ordering Finished")
print()

def input_due_in_names():
  for row in current_sheet.iter_rows(min_row=6, min_col=2, max_col=2, max_row=86):
    for cell in row:
      logging.debug("Cell value at " + str(cell.coordinate) + " is: " + str(cell.value))
      logging.debug("Cell data type is " + str(cell.data_type))
      debug_print = str(current_sheet[cell.coordinate])
      row_number = debug_print.strip("<Cell 'Camp Check'.B>")
      logging.debug("Row " + row_number)

      site_row = cell.row
      site_col = cell.column - 1
      site_coord = current_sheet.cell(site_row, site_col).value
      logging.debug("Site is: " + str(site_coord))
      cell_coordinate = str(cell.coordinate)
      logging.debug("cell_coordinate is " + cell_coordinate)
      
      try:
        logging.debug("Key found at " + names[str(site_coord)])
        logging.debug("Name: " + str(names[str(site_coord)]))

        logging.debug("Name Printing Works at " + str(cell.coordinate))

        # Enter name into spreadsheet
        current_sheet.cell(cell.row, cell.column, names[str(site_coord)])
        
        logging.debug("This cell value: " + str(current_sheet[cell.coordinate].value))
        
      except:
        logging.debug("No key found.")
        break

print("Beginning to input names")
logging.info("Beginning to input names")
print()
input_due_in_names()

# START DUE OUT ORDERING
print("Start Due Out Ordering")
logging.info("Start Due Out Ordering")

unsorted_dates = {}
dates = {}

for row in due_in_sheet.iter_rows(min_row=3, min_col=4, max_col=4, max_row=63):
  for cell in row:

    if str(cell.data_type) == "s":
      site_row = cell.row
      site_col = cell.column + 6
      site_coord = due_in_sheet.cell(site_row, site_col).value
      site_number = str(cell.value)
      logging.debug(str(site_number))
      
      print("Site #: " + str(cell.value) + " is due out on " + str(site_coord))
      logging.debug("Site #: " + str(cell.value) + " is due out on " + str(site_coord))
      
      unsorted_dates[site_number] = site_coord # changed from string

logging.debug("Printing unsorted dates")
logging.debug(str(unsorted_dates))

dates = dict(sorted(unsorted_dates.items(), key=lambda item: item[0]))

logging.debug("Printing sorted dates below...")
logging.debug(str(dates))
logging.debug("Printing list of keys in {dates}")
logging.debug(str(list(dates)))

def input_due_out_dates():
  for row in current_sheet.iter_rows(min_row=6, min_col=3, max_col=3, max_row=86):
    for cell in row:
      logging.debug("Cell value at " + str(cell.coordinate) + " is: " + str(cell.value))
      logging.debug("Cell data type is " + str(cell.data_type))
      debug_print = str(current_sheet[cell.coordinate])
      row_number = debug_print.strip("<Cell 'Camp Check'.B>")
      logging.debug("Row " + row_number)

      site_row = cell.row
      site_col = cell.column - 2
      site_coord = current_sheet.cell(site_row, site_col).value
      logging.debug("Site is: " + str(site_coord))
      cell_coordinate = str(cell.coordinate)
      logging.debug("cell_coordinate is " + cell_coordinate)
      
      try:
        logging.debug("Key found at " + str(dates[str(site_coord)]))
        logging.debug("Date: " + str(dates[str(site_coord)]))
        logging.debug("")

        logging.debug("Date Printing Works at " + str(cell.coordinate))

        # Enter date into spreadsheet
        formatted_date = dates[str(site_coord)]
        
        formatted_date = datetime.strptime(formatted_date, "%m/%d/%Y")
        logging.debug("formatted_date datatype: " + str(type(formatted_date)))
        

        logging.debug("formatted_date: " + str(formatted_date))
        current_sheet.cell(
                cell.row, 
                cell.column, 
                #str(dates[str(site_coord)][1:-5])
                formatted_date
                )
        
        logging.debug("This cell value: " + str(current_sheet[cell.coordinate].value))
        logging.debug("This cell datatype: " + str(current_sheet[cell.coordinate].data_type))
        
      except:
        logging.debug("No key found.")
        logging.debug("")
        break

# END DUE OUT ORDERING
input_due_out_dates()

# FINAL STEP
# Save the Spreadsheet!
print()
print("Saving camp check")
logging.info("Saving camp check")
camp_check.save(dst)
logging.info("Camp Check Complete")
input("Press any key to exit")
